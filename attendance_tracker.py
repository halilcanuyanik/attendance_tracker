import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk


def process_entry_gui(tcid, feedback_label, info_frame, root):
    personnel_df = pd.read_excel("staff_list.xlsx")
    tcid = str(tcid)
    date_today = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M")

    if not Path("attendance_log.xlsx").exists():
        attendance_df = pd.DataFrame(columns=[
            "Date", "ID Number", "First Name", "Last Name", "Department", "Entry Time", "Exit Time"
        ])
    else:
        attendance_df = pd.read_excel("attendance_log.xlsx", dtype={
            "ID Number": str,
            "Entry Time": str,
            "Exit Time": str
        })

    today_record = attendance_df[
        (attendance_df["Date"] == date_today) &
        (attendance_df["ID Number"] == tcid)
    ]

    is_registered = tcid in personnel_df["ID Number"].astype(str).values

    if is_registered:
        person = personnel_df[personnel_df["ID Number"].astype(str) == tcid].iloc[0]
        first_name, last_name, department = person["First Name"], person["Last Name"], person["Department"]
    else:
        first_name, last_name, department = "X", "X", "X"

    open_record = today_record[today_record["Exit Time"].isna() | (today_record["Exit Time"] == "")]

    if not open_record.empty:
        idx = open_record.index[-1]
        attendance_df.at[idx, "Exit Time"] = time_now
        action = "Exit"
    else:
        new_record = {
            "Date": date_today,
            "ID Number": tcid,
            "First Name": first_name,
            "Last Name": last_name,
            "Department": department,
            "Entry Time": time_now,
            "Exit Time": ""
        }
        attendance_df = pd.concat([attendance_df, pd.DataFrame([new_record])], ignore_index=True)
        action = "Entry"

    attendance_df.to_excel("attendance_log.xlsx", index=False)

    wb = load_workbook("attendance_log.xlsx")
    ws = wb.active
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20
        has_x = any(cell.value == "X" for cell in row)
        if has_x:
            for cell in row:
                cell.fill = red_fill

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 12

    wb.save("attendance_log.xlsx")

    # Update feedback
    feedback_label.config(
        text=f"{action} recorded at {time_now}",
        foreground="green" if action == "Entry" else "#007acc"
    )

    for widget in info_frame.winfo_children():
        widget.destroy()

    ttk.Label(info_frame, text=f"Name: {first_name} {last_name}", font=("Calibri", 15, "bold")).pack(anchor="w", pady=2)
    ttk.Label(info_frame, text=f"Department: {department}", font=("Calibri", 15)).pack(anchor="w", pady=2)
    ttk.Label(info_frame, text=f"{action} Time: {time_now}", font=("Calibri", 15)).pack(anchor="w", pady=2)

    # Auto-clear info after 5 seconds
    def clear_info():
        for widget in info_frame.winfo_children():
            widget.destroy()
        feedback_label.config(text="")

    root.after(5000, clear_info)


def launch_gui():
    root = tk.Tk()
    root.title("Attendance Tracker")
    root.geometry("700x500")
    root.resizable(False, False)

    style = ttk.Style()
    style.configure("TLabel", font=("Calibri", 13))
    style.configure("TButton", font=("Calibri", 13))
    style.configure("TEntry", font=("Calibri", 13))

    # Top frame with clock
    top_frame = ttk.Frame(root, padding=(20, 10))
    top_frame.pack(fill="x")

    ttk.Label(top_frame, text="Attendance Tracking System", font=("Calibri", 18, "bold")).pack(side="left")

    clock_label = ttk.Label(top_frame, font=("Calibri", 14))
    clock_label.pack(side="right")

    def update_clock():
        now = datetime.now().strftime("%H:%M:%S")
        clock_label.config(text=now)
        root.after(1000, update_clock)

    update_clock()

    # Main frame
    main_frame = ttk.Frame(root, padding=30)
    main_frame.pack(fill="both", expand=True)

    ttk.Label(main_frame, text="Scan or Enter ID Number:", font=("Calibri", 14)).pack(pady=(0, 10))

    id_entry = ttk.Entry(main_frame, width=35)
    id_entry.pack()

    feedback_label = ttk.Label(main_frame, text="", font=("Calibri", 13))
    feedback_label.pack(pady=10)

    info_frame = ttk.Frame(main_frame)
    info_frame.pack(pady=10, fill="x")

    def on_submit():
        tcid = id_entry.get().strip()
        if len(tcid) == 11 and tcid.isdigit():
            process_entry_gui(tcid, feedback_label, info_frame, root)
            id_entry.delete(0, tk.END)
        else:
            messagebox.showerror("Invalid Input", "ID number must be exactly 11 digits.")

    ttk.Button(main_frame, text="Submit", command=on_submit).pack(pady=10)

    id_entry.bind("<Return>", lambda event: on_submit())

    root.mainloop()


if __name__ == "__main__":
    launch_gui()
